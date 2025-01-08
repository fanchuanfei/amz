from amz_py.dao.SourceProduct import Product
from typing import Set
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 数据源文件转换
# folder_path：存放数据源文件夹的地址
def convert_csv_to_xlsx(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            csv_file = os.path.join(folder_path, filename)
            excel_file = os.path.splitext(csv_file)[0] + '.xlsx'
            df = pd.read_csv(csv_file)
            df.to_excel(excel_file, index=False)
            os.remove(csv_file)
            print(f'{filename} 已转换为 {os.path.basename(excel_file)} 并已删除原CSV文件')
        #


# 获取前一天的日期

def getpredate(folder_path):
    yesterday = datetime.now() - timedelta(days=1)
    formatted_yesterday = yesterday.strftime('%d-%m-%y')
    new_file_name = f"BusinessReport-{formatted_yesterday}.xlsx"

    # 拼接为文件
    fileaddress = os.path.join(folder_path, new_file_name)

# 打开资源表 装入字典

def getdate(fileaddress):
    sourceworkboook = load_workbook(fileaddress)
    sourceworksheet = sourceworkboook.active

    # key：父Asin  ， values:集合类型的子Sku，子Asin
    productdict: dict[str, Set[Product]] = {}

    # 把数据加入到字典中
    for i in range(2, sourceworksheet.max_row+1):
        parensasin = sourceworksheet.cell(i, 1).value


        if parensasin is None:       #数据读完直接退出
            break

        # 存储数据到定义的的变量



        childasin = sourceworksheet.cell(i, 2).value
        titlestr = sourceworksheet.cell(i, 3).value
        titlestr.split(",")
        title = titlestr[0]

        amdatetime = datetime.now()
        t_session = sourceworksheet.cell(i, 4).value
        t_pageviews = sourceworksheet.cell(i, 8).value
        t_cartpercentage = sourceworksheet.cell(i, 12).value
        t_sales = sourceworksheet.cell(i, 14).value
        t_conversionRate = sourceworksheet.cell(i, 16).value
        t_salesVolume = sourceworksheet.cell(i, 18).value
        t_orders =sourceworksheet.cell(i, 20).value



        product = Product()


        #把创建好的类放入字典的值中
        productchildset = {product}

        #判断是否为多Sku，如果是不用创建新集合，追加到对应父Asin下
        if parensasin == childasin:

            productdict.update({parensasin: productchildset})

        elif productdict.get(parensasin) is not None:

            productchildset = productdict.get(parensasin)

            productchildset.add(product)

        else:
            productdict.update({parensasin: productchildset})


    return productdict










