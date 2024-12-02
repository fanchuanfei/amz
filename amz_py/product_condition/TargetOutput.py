from openpyxl import load_workbook
from copy import deepcopy


def copy_cell(copy_from, paste_to_cell):
    """复制粘贴某个区域
    :param copy_from 复制源
    :param paste_to_cell 粘贴的左上角
    """
    # 记录边缘值
    for _copy_row in copy_from:  # 循环每一行
        for _row_cell in _copy_row:  # 循环每一列
            paste_to_cell.value = _row_cell.value
            paste_to_cell._style = deepcopy(_row_cell._style)  # 复制样式
            paste_to_cell = paste_to_cell.offset(row=0, column=1)  # 右移1格
        paste_to_cell = paste_to_cell.offset(row=1, column=-len(_copy_row))


# 1 代表已存在 ，0代表不存在
#判断一个工作博中有没有以parentasin为title的工作表
def if_exists(parentasin,targetworkbook):
    sheets = targetworkbook.sheetnames
    for sheet in sheets:
        if sheet.title == parentasin:
            return 1
    return 0



# 通过SourceAsin  判断表中是否有 TargetAsin（表中对映的工作表title） 没有就追加一个新的工作表
def update_sheet(productdict):
    targetworkbook = load_workbook(r"C:\Users\Administrator\Desktop\小蜗牛\产品基础数据.xlsx")
    for parentkey in productdict.keys():
        flag = if_exists(parentkey,targetworkbook)
        #新产品
        if flag == 0:
            #复制一个新的工作表
            newsheet = targetworkbook.copy_worksheet(targetworkbook["temp"])
            #录入数据行
            newsheet["B1"] = 4
            #录入Sku数量
            newsheet["D1"] = 1
            newsheet.title = parentkey

            #录入产品列数据

        #只有一个Sku的老产品
        if flag == 1 and len(productdict.get(parentkey)) == 1:
            print("只有一个数SKU老产品")


        if flag == 1 and len(productdict.get(parentkey)) < 1:
            print("多个SKU老产品")






        targetworkbook.save(r"C:\Users\Administrator\Desktop\小蜗牛\产品基础数据.xlsx")

















# 把每个产品信息遍历入表格中
# 多子Asin如何遍历到一个工作波中  步长为所需的长途，获取每个单元格 步长+相对位置

# def update_data(parentasin):
