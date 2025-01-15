import os

import openpyxl.utils
from openpyxl.reader.excel import load_workbook

from amz_py.logistics.restock import remove_colume

folder_path = r"C:\Users\Administrator\Desktop\test\1.xlsx"

sourceworkboook = load_workbook(folder_path)
sourceworksheet = sourceworkboook.active

sourceworksheet.cell(1,1,34)


my = sourceworksheet.cell(1,1).value

print(my)



