import os
from collections import OrderedDict
import openpyxl.utils
import pandas as pd
from openpyxl.reader.excel import load_workbook
from amz_py.logistics.dao.Salesandinventory import SalesInventory


# 将 txt文件转换为 xlsx文件
def convert_cvstxt_to_xlsx(folder_path):
    # Iterate through all files in the folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        # Check if it's a file
        if os.path.isfile(file_path):
            # Process .txt files
            if filename.endswith('.txt'):
                try:
                    data = pd.read_csv(file_path, sep='\t', engine='python')  # Assuming tab-delimited text files
                    output_file = os.path.join(folder_path, f"{os.path.splitext(filename)[0]}.xlsx")
                    data.to_excel(output_file, index=False)
                    print(f"Converted: {filename} -> {os.path.basename(output_file)}")
                except Exception as e:
                    print(f"Failed to convert {filename}: {e}")

            # Process .csv files
            elif filename.endswith('.csv'):
                try:
                    data = pd.read_csv(file_path)
                    output_file = os.path.join(folder_path, f"{os.path.splitext(filename)[0]}.xlsx")
                    data.to_excel(output_file, index=False)
                    print(f"Converted: {filename} -> {os.path.basename(output_file)}")
                except Exception as e:
                    print(f"Failed to convert {filename}: {e}")

# 读取库存数据
def read_inventory(folder_path):
    inventory_dict: dict = {}
    inventory_fileaddress = os.path.join(folder_path, "库存.xlsx")
    sourceworkboook = load_workbook(inventory_fileaddress)
    sourceworksheet = sourceworkboook.active

    for i in range(2, sourceworksheet.max_row+1):
        # 通过SKu判断是不是我的产品 ，不是直接下一个
        if sourceworksheet.cell(i,1).value.split("-")[0] != 3:
            continue


        asin = sourceworksheet.cell(i,3).value
        sku = sourceworksheet.cell(i,1).value
        product_name = sourceworksheet.cell(i,4).value
        fulfillable_quantity = sourceworksheet.cell(i,11).value   # 可售数量
        reserved_quantity = sourceworksheet.cell(i,13).value # 预留数量
        currentAndReserve_inventory = fulfillable_quantity +  reserved_quantity

        inbound_working_quantity = sourceworksheet.cell(i,16).value  #已发货，但在处理 。 例如没填追踪码
        inbound_shipped_quantity = sourceworksheet.cell(i,17).value #已发货，在途数量

        Inbound_inventory = inbound_working_quantity + inbound_shipped_quantity

        salesinventory = SalesInventory(asin, sku, product_name, currentAndReserve_inventory,Inbound_inventory)

        inventory_dict.update({sku:salesinventory})




    return inventory_dict

# 读取前七天的销量
def read_Sales(folder_path, inventory_dict):
    inventory_fileaddress = os.path.join(folder_path,"销量.xlsx")
    sourceworkboook = load_workbook(inventory_fileaddress)
    sourceworksheet = sourceworkboook.active

    for i in range(2, sourceworksheet.max_row+1):
        # 通过SKu判断是不是我的产品 ，不是直接下一个
        if sourceworksheet.cell(i, 1).value.split("-")[0] != 3:
            continue

        sku = sourceworksheet.cell(i,12).value

        quality = sourceworksheet.cell(i,15).value

        salesinventory = inventory_dict.get(sku)

        salesinventory.add_quality(quality)

    return inventory_dict

# 开始读取
def read_inventoryAndSales(folder_path):
    inventory_dict = read_inventory(folder_path)

    inventoryAndSales_dict = read_Sales(folder_path,inventory_dict)

    return inventoryAndSales_dict

# 移动指定行的数据 并判断有几行数据
def remove_colume(sourceworksheet, row):
    """
    将指定单元格中有数据的列向前移动一列。
    :param sourceworksheet: 表格
    :param row: 指定行号
    """
    # 用来记录有几列有数据

    count = 1

    # 将列字母转换为数字索引
    start_index = openpyxl.utils.cell.column_index_from_string('L')
    end_index = openpyxl.utils.cell.column_index_from_string('P')

    # 遍历列，从结束列向起始列逆向遍历
    for col in range(start_index, end_index+1):

        current_cell_value = sourceworksheet.cell(row, col).value

        if current_cell_value is None:  # 空的找下一列
            continue
        else:
            sourceworksheet.cell(row, col, current_cell_value)   # 不是空的把这列数据覆盖到前一列数据
            count += 1


    return count

# 获得增长率
def get_YoY_growth(row, sourceworksheet, count):
    """
    Calculate the year-over-year (YoY) growth rate for the given columns.

    Parameters:
    row (int): The row number to calculate the growth rate for.
    sourceworksheet (openpyxl worksheet object): The worksheet containing the data.
    count (int): The number of columns to consider, starting from column P.

    Returns:
    float: The YoY growth rate.
    """
    # Define the column letters
    columns = ['P', 'O', 'N', 'M', 'L', 'K']

    # Get the values for the current row
    current_values = [sourceworksheet[f'{col}{row}'].value for col in columns[:count]]

    # Get the values for the previous year (i.e., the row above)
    previous_values = [sourceworksheet[f'{col}{row-1}'].value for col in columns[:count]]

    # Calculate the sum of the current values
    current_sum = sum(current_values)

    # Calculate the sum of the previous values
    previous_sum = sum(previous_values)

    # Calculate the YoY growth rate
    growth_rate = ((current_sum - previous_sum) / previous_sum) * 100

    return growth_rate

# 将字典的数据排序
def sort_dict_values_by_key(input_dict):
    """
    根据字典的键对其值进行排序，并存入一个列表。
    参数：
        input_dict (dict): 输入的字典。
    返回：
        list: 按键排序后的值的列表。
    """
    # 使用字典按键排序
    sorted_dict = OrderedDict(sorted(input_dict.items()))

    # 提取排序后的值
    sorted_values = list(sorted_dict.values())

    return sorted_values

# 写入销量和库存
def write_Salesandinventory(restocking_address, inventoryAndSales_dict):


    # 1.打开库存表文件
    sourceworkboook = load_workbook(restocking_address)
    sourceworksheet = sourceworkboook['产品库存']

    # 2.依次读取表中的sku 判断表示是否有字典中的SKU


    for row in range(5, sourceworksheet.max_row+1):
        sku = sourceworksheet.cell(row, 2)

        # 3.如果字典中存在 更新表的数据 ，并删除字典中的数据
        if sku in inventoryAndSales_dict:
            # 获取库存对象
            salesInventory = inventoryAndSales_dict.get(sku)
            # 填入可售库存
            currentAndReserve_inventory = salesInventory.get_currentAndReserve_inventory()
            sourceworksheet.cell(row, 4, currentAndReserve_inventory)
            # 填入入库库存
            Inbound_inventory = salesInventory.get_Inbound_inventory()
            sourceworksheet.cell(row,6,Inbound_inventory)
            # 将过往周数据列向前移动 ,并得到有几列数据
            count = remove_colume(sourceworksheet,row)
            # 填写上一周的销量
            quality = salesInventory.get_quality()
            sourceworksheet.cell(row, 16, quality)
            # 填写增长率
            YoY_growth = get_YoY_growth(row,sourceworksheet,count)
            sourceworksheet.cell(row,18,YoY_growth)

            inventoryAndSales_dict.pop(sku)

    # 5.将剩余表中没有的数据，获取所有的key根据倒数第二个进行排序取出 遍历到表中

    sorted_values = sort_dict_values_by_key(inventoryAndSales_dict)

    count = 0
    for row in range(sourceworksheet.max_row+1,len(sorted_values)):
        salesInventory = sorted_values[count]

        sku = salesInventory.get_sku()
        sourceworksheet.cell(row,2,sku)

        product_name = salesInventory.get_product_name()
        sourceworksheet.cell(row, 3, product_name)

        currentAndReserve_inventory = salesInventory.get_currentAndReserve_inventory()
        sourceworksheet.cell(row, 4, currentAndReserve_inventory)

        Inbound_inventory = salesInventory.get_Inbound_inventory()
        sourceworksheet.cell(row, 5, Inbound_inventory)

        quality = salesInventory.get_quality()
        sourceworksheet.cell(row, 16, quality)

        sourceworksheet.cell(row, 8, 100)

        count += 1














